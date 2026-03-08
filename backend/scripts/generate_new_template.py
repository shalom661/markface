import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_dynamic_template(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matérias-Primas"

    # Only the bare minimum + 2 examples instead of all 45 columns
    headers = [
        "Categoria", 
        "Descricao", 
        "Unidade", 
        "Fornecedor", 
        "Codigo_Interno", 
        "Cor_do_Item", 
        "Composicao_do_Tecido",
        "Colecao_Personalizada",
        "Qualquer_Outra_Coluna_Aqui"
    ]

    ws.append(headers)

    # Styling headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

    # Example rows
    examples = [
        ["Tecido", "Malha 100% Algodão", "Metro", "Textil XYZ", "TEC001", "Preto", "100% Algodão", "Verão 2026", "Dado Extra"],
        ["Botão", "Botão Massa 4 Furos M", "Unidade", "Aviamentos ABC", "BOT002", "Transparente", "Acrílico", "Inverno", "Outro dado"],
        ["Zíper", "Zíper Invisível 15cm", "Unidade", "Ziper Cia", "ZIP003", "Branco", "Nylon", "Alto Inverno", "Nada"],
    ]

    for row in examples:
        ws.append(row)

    wb.save(output_path)
    print(f"Generated Excel at: {output_path}")

if __name__ == "__main__":
    root_file = r"d:\Shalom\Mark Face\MARK FACE HUB\markface\MODELO FLEXÍVEL MATÉRIA PRIMA.xlsx"
    backend_file = r"d:\Shalom\Mark Face\MARK FACE HUB\markface\backend\app\templates\template_materias_primas.xlsx"
    
    create_dynamic_template(root_file)
    create_dynamic_template(backend_file)
