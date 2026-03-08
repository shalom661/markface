"""
app/services/import_service.py
Service to parse dynamic Excel files and import raw materials into the database.
"""

from io import BytesIO
from typing import Any
import unicodedata
import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.raw_material import RawMaterial
from app.models.supplier import Supplier

# Fields that go directly to model columns. Everything else goes to category_fields JSONB.
MODEL_FIELDS = {"descricao", "categoria", "sub_categoria", "codigo_interno", "fornecedor", "unidade"}

# Required fields for validation
REQUIRED_FIELDS = ["categoria", "descricao", "unidade"]


def _clean(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _normalize_header(header: Any) -> str:
    """
    Normalizes a header string: 'Sub Categoria' -> 'sub_categoria'
    """
    if not header:
        return ""
    s = str(header).strip().lower()
    # Remove accents
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    # Replace common separators with underscores
    s = s.replace(' ', '_').replace('-', '_').replace('.', '')
    return s


async def _get_supplier_map(db: AsyncSession) -> dict[str, str]:
    result = await db.execute(select(Supplier).where(Supplier.active == True))
    suppliers = result.scalars().all()
    return {sup.name.lower().strip(): str(sup.id) for sup in suppliers}


async def import_raw_materials_from_excel(
    db: AsyncSession,
    file_contents: bytes,
) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_contents), data_only=True)

    # Try to find the 'Matérias-Primas' sheet; if not found use the first sheet
    if "Matérias-Primas" in wb.sheetnames:
        ws = wb["Matérias-Primas"]
    else:
        ws = wb.active

    supplier_map = await _get_supplier_map(db)

    created = 0
    errors = []
    skipped = 0

    header_row_idx = None
    header_map = {}  # col_index -> normalized_header_name

    # 1. Identify Header Row
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or all(v is None for v in row): 
            continue
            
        norm_row = [_normalize_header(c) for c in row]
        # We assume the row containing both 'categoria' and 'descricao' is the header row
        if "categoria" in norm_row or "descricao" in norm_row:
            header_row_idx = row_idx
            for i, col_name in enumerate(norm_row):
                if col_name:
                    header_map[i] = col_name
            break

    if not header_row_idx or not header_map:
        return {
            "message": "Falha na importação",
            "created": 0,
            "skipped": 0,
            "errors": [{"row": 0, "error": "Cabeçalho não encontrado. Certifique-se de que a planilha possui colunas como 'Categoria' e 'Descricao'.", "data": {}}]
        }

    # 2. Process Data Rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            skipped += 1
            continue

        # Map values to dynamic keys based on the header map
        row_data: dict[str, str | None] = {}
        for i, val in enumerate(row):
            if i in header_map:
                row_data[header_map[i]] = _clean(val)

        # Check required fields
        missing = [f for f in REQUIRED_FIELDS if not row_data.get(f)]
        if missing:
            errors.append({
                "row": row_idx,
                "error": f"Campos obrigatórios faltando: {', '.join(missing)}",
                "data": {k: v for k, v in row_data.items() if v},
            })
            continue

        # Resolve supplier
        supplier_id = None
        sup_name = row_data.get("fornecedor")
        if sup_name and sup_name.lower().strip() in supplier_map:
            supplier_id = supplier_map[sup_name.lower().strip()]

        # Build category fields from all non-model, non-empty fields
        category_fields: dict[str, Any] = {}
        for key, val in row_data.items():
            if key in MODEL_FIELDS or not val:
                continue
            category_fields[key] = val

        # Create the RawMaterial record
        material = RawMaterial(
            description=row_data.get("descricao", ""),
            internal_code=row_data.get("codigo_interno"),
            category=row_data.get("categoria", ""),
            subcategory=row_data.get("sub_categoria"),
            unit=row_data.get("unidade", ""),
            supplier_id=supplier_id,
            category_fields=category_fields,
            active=True,
        )
        db.add(material)
        created += 1

    await db.commit()

    return {
        "message": f"Importação concluída: {created} itens criados, {len(errors)} erros, {skipped} linhas ignoradas.",
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }
