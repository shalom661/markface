from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook
from app.core.deps import get_db
from app.schemas.finance import FixedCostCreate, FixedCostRead, PurchaseCreate, PurchaseRead, SalesModalityCreate, SalesModalityRead
from app.services import finance_service

router = APIRouter()

@router.get("/fixed-costs", response_model=list[FixedCostRead])
async def get_fixed_costs(db: AsyncSession = Depends(get_db)):
    return await finance_service.list_fixed_costs(db)

@router.post("/fixed-costs", response_model=FixedCostRead)
async def post_fixed_cost(schema: FixedCostCreate, db: AsyncSession = Depends(get_db)):
    return await finance_service.create_fixed_cost(db, schema)

@router.delete("/fixed-costs/{cost_id}")
async def remove_fixed_cost(cost_id: str, db: AsyncSession = Depends(get_db)):
    await finance_service.delete_fixed_cost(db, cost_id)
    return {"message": "Deleted"}

@router.get("/purchases", response_model=list[PurchaseRead])
async def get_purchases(db: AsyncSession = Depends(get_db)):
    return await finance_service.list_purchases(db)

@router.post("/purchases", response_model=PurchaseRead)
async def post_purchase(schema: PurchaseCreate, db: AsyncSession = Depends(get_db)):
    try:
        return await finance_service.create_purchase(db, schema)
    except Exception as e:
        # Check if it's a known error or a generic one
        detail = str(e)
        if "ForeignKey" in detail or "foreign key" in detail.lower():
            detail = "Erro de integridade: Fornecedor ou Item não encontrado."
        raise HTTPException(status_code=500, detail=detail)

@router.get("/sales-modalities", response_model=list[SalesModalityRead])
async def get_sales_modalities(db: AsyncSession = Depends(get_db)):
    return await finance_service.list_sales_modalities(db)

@router.post("/sales-modalities", response_model=SalesModalityRead)
async def post_sales_modality(schema: SalesModalityCreate, db: AsyncSession = Depends(get_db)):
    return await finance_service.create_sales_modality(db, schema)

@router.delete("/sales-modalities/{modality_id}")
async def remove_sales_modality(modality_id: str, db: AsyncSession = Depends(get_db)):
    await finance_service.delete_sales_modality(db, modality_id)
    return {"message": "Deleted"}

@router.get("/finance/export-costs/{modality_id}")
async def export_costs(modality_id: str, db: AsyncSession = Depends(get_db)):
    data = await finance_service.get_detailed_variant_costs(db, modality_id)
    
    if not data:
        raise HTTPException(status_code=404, detail="Nenhum dado encontrado para esta modalidade.")

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Custos"
    
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subtotal_font = Font(bold=True)
    money_format = '"R$ "#,##0.00'
    
    # Headers
    headers = ["Produto/Item", "SKU", "Qtd", "Unidade", "Vlr Unit. (Médio)", "Subtotal Item", "Preço Yield"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    current_row = 2
    for p in data:
        # Product Header with bold name
        ws.cell(row=current_row, column=1, value=p["product_name"]).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=p["sku"]).font = Font(bold=True)
        current_row += 1
        
        # Materials
        for m in p["materials"]:
            ws.cell(row=current_row, column=1, value=f"  ↳ {m['material_name']}")
            ws.cell(row=current_row, column=3, value=float(m["quantity"]))
            ws.cell(row=current_row, column=4, value=m["unit"])
            ws.cell(row=current_row, column=5, value=float(m["avg_price"])).number_format = money_format
            ws.cell(row=current_row, column=6, value=float(m["item_cost"])).number_format = money_format
            current_row += 1
        
        # Fixed Share
        if p["fixed_share"] > 0:
            ws.cell(row=current_row, column=1, value="  ↳ Rateio Gastos Fixos")
            ws.cell(row=current_row, column=6, value=float(p["fixed_share"])).number_format = money_format
            current_row += 1
            
        # Modalidade Info (Taxes and Fees)
        ws.cell(row=current_row, column=1, value=f"  ↳ Taxa {p['modality_name']} ({p['tax_rate']}%)")
        # Formula for tax value: Yield * Rate%
        ws.cell(row=current_row, column=6, value=float(p["yield_price"] * (p["tax_rate"] / 100))).number_format = money_format
        current_row += 1
        
        if p["fixed_fee"] > 0 or p["extra_cost"] > 0:
            ws.cell(row=current_row, column=1, value="  ↳ Taxas Fixas/Extras Modalidade")
            ws.cell(row=current_row, column=6, value=float(p["fixed_fee"] + p["extra_cost"])).number_format = money_format
            current_row += 1

        # Product Summary Row
        ws.cell(row=current_row, column=1, value=f"TOTAL {p['product_name']}").font = subtotal_font
        ws.cell(row=current_row, column=6, value=float(p["base_cost"])).font = subtotal_font
        ws.cell(row=current_row, column=6).number_format = money_format
        
        # Final Yield Price highlighted
        yield_cell = ws.cell(row=current_row, column=7, value=float(p["yield_price"]))
        yield_cell.font = Font(bold=True, color="0070C0") # Blue bold
        yield_cell.number_format = money_format
        
        # Add a separator blank row
        current_row += 2

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 5

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"exportacao_custos_{p['modality_name'].replace(' ', '_')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
