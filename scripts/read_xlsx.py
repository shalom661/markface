import openpyxl

wb = openpyxl.load_workbook(r"c:\MARK FACE HUB\EXEMPLO MATÉRIA PRIMA.xlsx")
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Only show first 20 rows
        if row_idx > 20:
            print("  ...(truncated)")
            break
        # Only print row if it has any non-None value
        vals = [str(v) if v is not None else "" for v in row]
        if any(v.strip() for v in vals):
            print(f"  Row {row_idx}: {vals}")
